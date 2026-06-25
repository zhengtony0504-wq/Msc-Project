#!/usr/bin/env python3
"""
Build clean_panel_v1.csv from a Bloomberg BDH export where each ticker block
has its own visible Date column.

Expected input layout:
- One sheet named "BDH_Dates_Visible" or the first sheet.
- Start date in B1, end date in B2.
- Repeated blocks of 11 columns: Dates + 10 Bloomberg fields.
- Ticker name is in row 5, first field column of each block.
- Field names are in row 7.
- Data starts at row 8 after Bloomberg refresh.

Usage:
    python build_clean_panel_from_date_visible_export_v2.py refreshed_export.xlsx clean_panel_v1.csv

The output:
    date, source_date, year_month, ticker, <fields...>, next_month_return
"""
from __future__ import annotations

import sys
from pathlib import Path
import datetime as dt
import numpy as np
import pandas as pd
import openpyxl


N_FIELDS = 10
BLOCK_WIDTH = 11
TICKER_ROW = 5
HEADER_ROW = 7
DATA_START_ROW = 8


def clean_value(x):
    if x is None:
        return np.nan
    if isinstance(x, str):
        xs = x.strip()
        if xs == "" or xs.startswith("#N/A") or xs in {"#VALUE!", "#REF!", "#DIV/0!", "#NAME?", "#NULL!", "#NUM!"}:
            return np.nan
    return x


def to_timestamp_or_none(x):
    if isinstance(x, dt.datetime):
        return pd.Timestamp(x.date())
    if isinstance(x, dt.date):
        return pd.Timestamp(x)
    return None


def main() -> int:
    if len(sys.argv) < 2:
        print("Usage: python build_clean_panel_from_date_visible_export.py refreshed_export.xlsx [clean_panel_v1.csv]", file=sys.stderr)
        return 2

    input_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2]) if len(sys.argv) >= 3 else Path("clean_panel_v1.csv")
    audit_path = output_path.with_name(output_path.stem + "_audit.md")

    if not input_path.exists():
        print(f"Input file not found: {input_path}", file=sys.stderr)
        return 2

    # Use read_only=False because some generated xlsx files do not store a <dimension> tag.
    # In read_only mode openpyxl may then report max_row/max_column as None.
    # Normal mode calculates dimensions from populated cells and is still fine for this workbook size.
    wb = openpyxl.load_workbook(input_path, data_only=True, read_only=False)
    ws = wb["BDH_Dates_Visible"] if "BDH_Dates_Visible" in wb.sheetnames else wb[wb.sheetnames[0]]

    max_col = ws.max_column
    max_row = ws.max_row
    if max_col is None or max_row is None:
        dim = ws.calculate_dimension()
        if dim and dim != "A1:A1":
            from openpyxl.utils.cell import range_boundaries
            _, _, max_col, max_row = range_boundaries(dim)
    if max_col is None or max_row is None:
        print("Could not determine worksheet dimensions. Open, refresh and save the workbook in Excel, then rerun.", file=sys.stderr)
        return 1

    records = []
    tickers = []
    block_start = 1
    while block_start <= max_col:
        ticker = ws.cell(row=TICKER_ROW, column=block_start + 1).value
        if ticker is None or str(ticker).strip() == "":
            # Stop only after a long blank tail; otherwise advance.
            block_start += BLOCK_WIDTH
            continue

        ticker = str(ticker).strip()
        tickers.append(ticker)
        fields = [ws.cell(row=HEADER_ROW, column=block_start + 1 + i).value for i in range(N_FIELDS)]
        if any(f is None for f in fields):
            print(f"Warning: missing field header in block starting column {block_start} ({ticker})", file=sys.stderr)

        for row_idx in range(DATA_START_ROW, max_row + 1):
            source_date = to_timestamp_or_none(ws.cell(row=row_idx, column=block_start).value)
            if source_date is None:
                continue

            month_end = source_date + pd.offsets.MonthEnd(0)
            rec = {
                "date": month_end.date().isoformat(),
                "source_date": source_date.date().isoformat(),
                "year_month": month_end.strftime("%Y-%m"),
                "ticker": ticker,
            }
            any_value = False
            for i, field in enumerate(fields):
                val = clean_value(ws.cell(row=row_idx, column=block_start + 1 + i).value)
                rec[str(field)] = val
                if not (isinstance(val, float) and np.isnan(val)):
                    any_value = True
            if any_value:
                records.append(rec)

        block_start += BLOCK_WIDTH

    if not records:
        print("No records found. Has the Bloomberg workbook been refreshed and saved with values?", file=sys.stderr)
        return 1

    panel = pd.DataFrame.from_records(records)

    # Convert numeric fields where possible.
    metadata_cols = {"date", "source_date", "year_month", "ticker"}
    value_cols = [c for c in panel.columns if c not in metadata_cols]
    for col in value_cols:
        panel[col] = pd.to_numeric(panel[col], errors="coerce")

    # One ticker-month observation: if Bloomberg returns both trading month-end and calendar month-end,
    # keep the latest source_date within that calendar month for that ticker.
    panel = panel.sort_values(["ticker", "date", "source_date"])
    panel = panel.groupby(["ticker", "date"], as_index=False, sort=False).tail(1).copy()

    # Target variable: next-month stock return based only on PX_LAST.
    panel = panel.sort_values(["ticker", "date"]).reset_index(drop=True)
    if "PX_LAST" in panel.columns:
        panel["next_month_return"] = panel.groupby("ticker")["PX_LAST"].shift(-1) / panel["PX_LAST"] - 1
    else:
        panel["next_month_return"] = np.nan

    # Final sorted panel is easier to inspect by date.
    panel = panel.sort_values(["date", "ticker"]).reset_index(drop=True)
    panel.to_csv(output_path, index=False)

    missing_rates = panel.isna().mean().sort_values(ascending=False)
    ticker_obs = panel.groupby("ticker").size()
    audit = []
    audit.append("# clean_panel_v1 audit")
    audit.append("")
    audit.append(f"- Input workbook: `{input_path.name}`")
    audit.append(f"- Output CSV: `{output_path.name}`")
    audit.append(f"- Tickers detected: {panel['ticker'].nunique():,}")
    audit.append(f"- Calendar months detected: {panel['date'].nunique():,}")
    audit.append(f"- Rows: {len(panel):,}")
    audit.append(f"- Date range: {panel['date'].min()} to {panel['date'].max()}")
    audit.append(f"- Non-null next_month_return: {panel['next_month_return'].notna().sum():,}")
    audit.append(f"- Observations per ticker: min {ticker_obs.min():,}, median {int(ticker_obs.median()):,}, max {ticker_obs.max():,}")
    audit.append("")
    audit.append("## Missing rate by column")
    audit.append("")
    for col, rate in missing_rates.items():
        audit.append(f"- {col}: {rate:.2%}")
    audit_path.write_text("\n".join(audit), encoding="utf-8")

    print(f"Wrote {output_path} ({len(panel):,} rows)")
    print(f"Wrote {audit_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
